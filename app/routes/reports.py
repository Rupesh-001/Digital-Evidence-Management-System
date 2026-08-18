"""EvidenceChain topic-specific reporting system."""

import io
import re

from datetime import datetime, timezone
from xml.sax.saxutils import escape

from flask import (
    Blueprint,
    render_template,
    current_app,
    request,
    abort,
    send_file
)

from flask_login import (
    login_required,
    current_user
)

from app.models.evidence import (
    Case,
    Evidence,
    CustodyLog,
    AuditLog
)


reports_bp = Blueprint(
    "reports",
    __name__,
    url_prefix="/reports"
)


# ──────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────

def _db():
    return current_app.db


def _clean(value, default="—"):

    if value is None:
        return default

    text = str(value).strip()

    return text if text else default


def _fmt_dt(value, include_time=True):

    if not value:
        return "—"

    if hasattr(value, "strftime"):

        if include_time:
            return value.strftime(
                "%B %d, %Y %H:%M UTC"
            )

        return value.strftime(
            "%B %d, %Y"
        )

    return str(value)


def _safe_filename(value):

    value = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        str(value)
    )

    return (
        value.strip("._")
        or "report"
    )


# ──────────────────────────────────────────────────────────────
# Automatic descriptions
# ──────────────────────────────────────────────────────────────

def generate_case_description(
    case,
    evidence_count
):

    return (
        f"This report documents case "
        f"{_clean(case.get('case_id'))}, titled "
        f"'{_clean(case.get('title'))}'. "
        f"The case is classified as "
        f"{_clean(case.get('type'))} and currently has "
        f"a status of {_clean(case.get('status'))}. "
        f"The assigned priority is "
        f"{_clean(case.get('priority'))}. "
        f"The lead investigator is "
        f"{_clean(case.get('lead'))}. "
        f"The recorded case location is "
        f"{_clean(case.get('location'))}. "
        f"There are {evidence_count} evidence item(s) "
        f"linked to this case. "
        f"The recorded case description states: "
        f"{_clean(case.get('description'), 'No case description was recorded.')}"
    )


def generate_evidence_description(
    ev,
    case
):

    case_text = "No linked case record was found."

    if case:

        case_text = (
            f"It is associated with case "
            f"{_clean(case.get('case_id'))}, "
            f"'{_clean(case.get('title'))}'."
        )

    return (
        f"Evidence item "
        f"{_clean(ev.get('evidence_id'))} "
        f"is identified as "
        f"'{_clean(ev.get('name'))}' "
        f"and is classified as "
        f"{_clean(ev.get('type'))}. "
        f"The current evidence status is "
        f"{_clean(ev.get('status'))}. "
        f"{case_text} "
        f"The item was collected by "
        f"{_clean(ev.get('collected_by'))} "
        f"at {_clean(ev.get('location'))} "
        f"on {_fmt_dt(ev.get('collected_at'))}. "
        f"The recorded SHA-256 value is maintained "
        f"as the integrity fingerprint for this "
        f"evidence item. "
        f"Collection notes: "
        f"{_clean(ev.get('notes'), 'No additional collection notes were recorded.')}"
    )


def generate_custody_summary(
    ev,
    logs
):

    evidence_id = _clean(
        ev.get("evidence_id")
    )

    if not logs:

        return (
            f"No chain-of-custody entries are "
            f"currently recorded for evidence "
            f"{evidence_id}."
        )

    first = logs[0]
    last = logs[-1]

    return (
        f"The chain of custody for evidence "
        f"{evidence_id} contains "
        f"{len(logs)} recorded event(s). "
        f"The first recorded event was "
        f"'{_clean(first.get('action'))}' "
        f"on {_fmt_dt(first.get('timestamp'))}. "
        f"The latest recorded event was "
        f"'{_clean(last.get('action'))}' "
        f"on {_fmt_dt(last.get('timestamp'))}. "
        f"The latest recorded integrity state is "
        f"{_clean(last.get('integrity'))}."
    )


def generate_audit_summary(audits):

    if not audits:

        return (
            "No audit events are currently "
            "recorded in the system."
        )

    actions = {}

    for audit in audits:

        action = _clean(
            audit.get("action"),
            "Unknown"
        )

        actions[action] = (
            actions.get(action, 0) + 1
        )

    action_text = ", ".join(
        f"{action} ({count})"
        for action, count
        in sorted(actions.items())
    )

    return (
        f"This audit report contains "
        f"{len(audits)} recorded system event(s). "
        f"The recorded action distribution is: "
        f"{action_text}. "
        f"The report is generated from the "
        f"EvidenceChain audit collection and "
        f"does not modify the underlying records."
    )


# ──────────────────────────────────────────────────────────────
# HTML REPORTS
# ──────────────────────────────────────────────────────────────

@reports_bp.route("/")
@login_required
def generate_report():

    db = _db()

    cases = Case.all(db)
    evidence = Evidence.all(db)

    return render_template(
        "reports/generate_report.html",
        cases=cases,
        evidence=evidence
    )


@reports_bp.route("/case/<case_id>")
@login_required
def case_report(case_id):

    db = _db()

    case = Case.get(
        db,
        case_id
    )

    if not case:
        abort(404)

    evidence = Evidence.by_case(
        db,
        case_id
    )

    description = generate_case_description(
        case,
        len(evidence)
    )

    return render_template(
        "reports/case_report.html",
        case=case,
        evidence=evidence,
        generated_description=description
    )


@reports_bp.route("/evidence/<ev_id>")
@login_required
def evidence_report(ev_id):

    db = _db()

    ev = Evidence.get(
        db,
        ev_id
    )

    if not ev:
        abort(404)

    case = Case.get(
        db,
        ev.get("case_id", "")
    )

    custody = CustodyLog.by_evidence(
        db,
        ev_id
    )

    description = generate_evidence_description(
        ev,
        case
    )

    custody_summary = generate_custody_summary(
        ev,
        custody
    )

    return render_template(
        "reports/evidence_report.html",
        ev=ev,
        case=case,
        custody=custody,
        generated_description=description,
        custody_summary=custody_summary
    )


@reports_bp.route("/custody")
@login_required
def custody_report():

    db = _db()

    ev_id = request.args.get(
        "ev_id",
        ""
    ).strip()

    ev = (
        Evidence.get(db, ev_id)
        if ev_id
        else None
    )

    if ev_id and not ev:
        abort(404)

    logs = (
        CustodyLog.by_evidence(
            db,
            ev_id
        )
        if ev_id
        else []
    )

    evidence = Evidence.all(db)

    summary = (
        generate_custody_summary(
            ev,
            logs
        )
        if ev
        else None
    )

    return render_template(
        "reports/custody_report.html",
        ev=ev,
        logs=logs,
        evidence=evidence,
        ev_id=ev_id,
        summary=summary
    )


@reports_bp.route("/audit")
@login_required
def audit_report():

    db = _db()

    audits = AuditLog.all(
        db,
        limit=1000
    )

    summary = generate_audit_summary(
        audits
    )

    return render_template(
        "reports/audit_report.html",
        audits=audits,
        summary=summary
    )


# ──────────────────────────────────────────────────────────────
# PDF
# ──────────────────────────────────────────────────────────────

def _build_pdf(
    report_type,
    db,
    now_str
):

    try:

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import (
            getSampleStyleSheet,
            ParagraphStyle
        )
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            Spacer,
            Table,
            TableStyle
        )

    except ImportError:

        abort(
            500,
            "ReportLab is not installed."
        )

    styles = getSampleStyleSheet()

    elements = []

    brand_style = ParagraphStyle(
        "Brand",
        parent=styles["Heading1"],
        fontSize=22,
        leading=26,
        alignment=TA_CENTER,
        textColor=colors.HexColor(
            "#0f172a"
        ),
        spaceAfter=5
    )

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading2"],
        fontSize=16,
        leading=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor(
            "#2563eb"
        ),
        spaceAfter=5
    )

    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.HexColor(
            "#64748b"
        ),
        spaceAfter=12
    )

    heading_style = ParagraphStyle(
        "Heading",
        parent=styles["Heading3"],
        fontSize=12,
        leading=15,
        textColor=colors.HexColor(
            "#0f172a"
        ),
        spaceBefore=10,
        spaceAfter=6
    )

    body_style = ParagraphStyle(
        "Body",
        parent=styles["BodyText"],
        fontSize=9.5,
        leading=14,
        textColor=colors.HexColor(
            "#1e293b"
        ),
        spaceAfter=8
    )

    table_header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["BodyText"],
        fontSize=7.5,
        leading=9,
        textColor=colors.white,
        fontName="Helvetica-Bold"
    )

    table_cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["BodyText"],
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor(
            "#1e293b"
        )
    )

    def P(
        value,
        style=body_style
    ):

        text = escape(
            _clean(value)
        ).replace(
            "\n",
            "<br/>"
        )

        return Paragraph(
            text,
            style
        )

    def add_header(
        title,
        subject
    ):

        elements.append(
            Paragraph(
                "EvidenceChain",
                brand_style
            )
        )

        elements.append(
            Paragraph(
                escape(title),
                title_style
            )
        )

        elements.append(
            Paragraph(
                escape(subject),
                subtitle_style
            )
        )

        elements.append(
            Spacer(
                1,
                0.2 * cm
            )
        )

    def add_section(title):

        elements.append(
            Paragraph(
                escape(title),
                heading_style
            )
        )

    def add_text(text):

        elements.append(
            P(
                text,
                body_style
            )
        )

    def add_key_value(rows):

        data = []

        label_style = ParagraphStyle(
            "Label",
            parent=table_cell_style,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor(
                "#475569"
            )
        )

        for label, value in rows:

            data.append([
                P(
                    label,
                    label_style
                ),
                P(
                    value,
                    table_cell_style
                )
            ])

        table = Table(
            data,
            colWidths=[
                4.3 * cm,
                12.7 * cm
            ]
        )

        table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor(
                        "#f1f5f9"
                    )
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor(
                        "#cbd5e1"
                    )
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                )
            ])
        )

        elements.append(table)

        elements.append(
            Spacer(
                1,
                0.25 * cm
            )
        )

    def add_table(
        headers,
        rows,
        widths
    ):

        data = [
            [
                P(
                    h,
                    table_header_style
                )
                for h in headers
            ]
        ]

        for row in rows:

            data.append([
                P(
                    value,
                    table_cell_style
                )
                for value in row
            ])

        table = Table(
            data,
            colWidths=widths,
            repeatRows=1
        )

        table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#0f172a"
                    )
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor(
                            "#f8fafc"
                        )
                    ]
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor(
                        "#cbd5e1"
                    )
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                )
            ])
        )

        elements.append(table)

        elements.append(
            Spacer(
                1,
                0.25 * cm
            )
        )

    # ──────────────────────────────────────────────────────────
    # CASE REPORT
    # ──────────────────────────────────────────────────────────

    if report_type == "cases":

        case_id = request.args.get(
            "case_id",
            ""
        ).strip()

        if not case_id:

            abort(
                400,
                "case_id is required."
            )

        case = Case.get(
            db,
            case_id
        )

        if not case:

            abort(
                404,
                "Case not found."
            )

        evidence = Evidence.by_case(
            db,
            case_id
        )

        add_header(
            "Digital Case Report",
            (
                f"Case {case_id} | "
                f"Generated {now_str} by "
                f"{current_user.name}"
            )
        )

        add_section(
            "1. Case Identification"
        )

        add_key_value([
            (
                "Case ID",
                case.get("case_id")
            ),
            (
                "Case Title",
                case.get("title")
            ),
            (
                "Case Type",
                case.get("type")
            ),
            (
                "Status",
                case.get("status")
            ),
            (
                "Priority",
                case.get("priority")
            ),
            (
                "Lead Investigator",
                case.get("lead")
            ),
            (
                "Location",
                case.get("location")
            ),
            (
                "Opened",
                _fmt_dt(
                    case.get("created")
                )
            )
        ])

        add_section(
            "2. Automatically Generated Case Description"
        )

        add_text(
            generate_case_description(
                case,
                len(evidence)
            )
        )

        add_section(
            "3. Evidence Associated With This Case"
        )

        if evidence:

            add_table(
                [
                    "Evidence ID",
                    "Name",
                    "Type",
                    "Status",
                    "Collected By",
                    "Collected"
                ],
                [
                    [
                        e.get("evidence_id"),
                        e.get("name"),
                        e.get("type"),
                        e.get("status"),
                        e.get("collected_by"),
                        _fmt_dt(
                            e.get("collected_at"),
                            False
                        )
                    ]
                    for e in evidence
                ],
                [
                    2.5 * cm,
                    4.0 * cm,
                    2.2 * cm,
                    2.2 * cm,
                    3.0 * cm,
                    2.6 * cm
                ]
            )

        else:

            add_text(
                "No evidence is currently linked to this case."
            )

        add_section(
            "4. Case Conclusion"
        )

        add_text(
            f"This report is restricted to case "
            f"{case_id} and the evidence directly "
            f"linked to this case. No unrelated case "
            f"records are included."
        )

        subject = case_id

    # ──────────────────────────────────────────────────────────
    # EVIDENCE REPORT
    # ──────────────────────────────────────────────────────────

    elif report_type == "evidence":

        ev_id = request.args.get(
            "ev_id",
            ""
        ).strip()

        if not ev_id:

            abort(
                400,
                "ev_id is required."
            )

        ev = Evidence.get(
            db,
            ev_id
        )

        if not ev:

            abort(
                404,
                "Evidence not found."
            )

        case = Case.get(
            db,
            ev.get("case_id", "")
        )

        custody = CustodyLog.by_evidence(
            db,
            ev_id
        )

        add_header(
            "Digital Evidence Report",
            (
                f"Evidence {ev_id} | "
                f"Generated {now_str} by "
                f"{current_user.name}"
            )
        )

        add_section(
            "1. Evidence Identification"
        )

        add_key_value([
            (
                "Evidence ID",
                ev.get("evidence_id")
            ),
            (
                "Evidence Name",
                ev.get("name")
            ),
            (
                "Evidence Type",
                ev.get("type")
            ),
            (
                "Status",
                ev.get("status")
            ),
            (
                "Linked Case",
                ev.get("case_id")
            ),
            (
                "Case Title",
                case.get("title")
                if case
                else "Case record not found"
            )
        ])

        add_section(
            "2. Automatically Generated Evidence Description"
        )

        add_text(
            generate_evidence_description(
                ev,
                case
            )
        )

        add_section(
            "3. File and Collection Information"
        )

        add_key_value([
            (
                "Original Filename",
                ev.get("filename")
            ),
            (
                "File Size",
                ev.get("file_size")
            ),
            (
                "Storage Location",
                ev.get("location")
            ),
            (
                "Collected By",
                ev.get("collected_by")
            ),
            (
                "Collected At",
                _fmt_dt(
                    ev.get("collected_at")
                )
            ),
            (
                "Integrity Verified At",
                _fmt_dt(
                    ev.get(
                        "integrity_verified_at"
                    )
                )
            )
        ])

        add_section(
            "4. SHA-256 Integrity Verification"
        )

        verification_state = (
            "Verified"
            if ev.get("status") == "Verified"
            else
            "Compromised"
            if ev.get("status") == "Compromised"
            else
            "Not currently verified"
        )

        add_key_value([
            (
                "Stored SHA-256",
                ev.get("hash")
            ),
            (
                "Evidence Status",
                ev.get("status")
            ),
            (
                "Verification State",
                verification_state
            )
        ])

        add_section(
            "5. Collection Notes"
        )

        add_text(
            _clean(
                ev.get("notes"),
                "No collection notes were recorded."
            )
        )

        add_section(
            "6. Chain of Custody for This Evidence"
        )

        if custody:

            add_table(
                [
                    "Timestamp",
                    "Action",
                    "From",
                    "To",
                    "Handler",
                    "Integrity"
                ],
                [
                    [
                        _fmt_dt(
                            c.get("timestamp")
                        ),
                        c.get("action"),
                        c.get("from_party"),
                        c.get("to_party"),
                        c.get("handler"),
                        c.get("integrity")
                    ]
                    for c in custody
                ],
                [
                    2.8 * cm,
                    2.0 * cm,
                    2.8 * cm,
                    2.8 * cm,
                    3.0 * cm,
                    2.6 * cm
                ]
            )

        else:

            add_text(
                "No custody records are currently recorded."
            )

        add_section(
            "7. Automatically Generated Forensic Summary"
        )

        add_text(
            generate_custody_summary(
                ev,
                custody
            )
        )

        add_text(
            f"This report is restricted to evidence "
            f"{ev_id}. Only the selected evidence "
            f"record, its integrity information, "
            f"collection information, and its custody "
            f"records are included."
        )

        subject = ev_id

    # ──────────────────────────────────────────────────────────
    # CUSTODY REPORT
    # ──────────────────────────────────────────────────────────

    elif report_type == "custody":

        ev_id = request.args.get(
            "ev_id",
            ""
        ).strip()

        if not ev_id:

            abort(
                400,
                "ev_id is required."
            )

        ev = Evidence.get(
            db,
            ev_id
        )

        if not ev:

            abort(
                404,
                "Evidence not found."
            )

        logs = CustodyLog.by_evidence(
            db,
            ev_id
        )

        case = Case.get(
            db,
            ev.get("case_id", "")
        )

        add_header(
            "Chain of Custody Report",
            (
                f"Evidence {ev_id} | "
                f"Generated {now_str} by "
                f"{current_user.name}"
            )
        )

        add_section(
            "1. Evidence Reference"
        )

        add_key_value([
            (
                "Evidence ID",
                ev.get("evidence_id")
            ),
            (
                "Evidence Name",
                ev.get("name")
            ),
            (
                "Evidence Type",
                ev.get("type")
            ),
            (
                "Evidence Status",
                ev.get("status")
            ),
            (
                "Linked Case",
                ev.get("case_id")
            ),
            (
                "Case Title",
                case.get("title")
                if case
                else "Case record not found"
            ),
            (
                "SHA-256",
                ev.get("hash")
            )
        ])

        add_section(
            "2. Automatically Generated Custody Summary"
        )

        add_text(
            generate_custody_summary(
                ev,
                logs
            )
        )

        add_section(
            "3. Custody Timeline"
        )

        if logs:

            add_table(
                [
                    "Timestamp",
                    "Action",
                    "From",
                    "To",
                    "Handler",
                    "Reason",
                    "Location",
                    "Integrity"
                ],
                [
                    [
                        _fmt_dt(
                            c.get("timestamp")
                        ),
                        c.get("action"),
                        c.get("from_party"),
                        c.get("to_party"),
                        c.get("handler"),
                        c.get("reason"),
                        c.get("location"),
                        c.get("integrity")
                    ]
                    for c in logs
                ],
                [
                    2.3 * cm,
                    1.8 * cm,
                    2.2 * cm,
                    2.2 * cm,
                    2.4 * cm,
                    2.6 * cm,
                    2.3 * cm,
                    1.7 * cm
                ]
            )

        else:

            add_text(
                "No custody events are recorded."
            )

        add_section(
            "4. Custody Conclusion"
        )

        add_text(
            f"This report contains only the chain "
            f"of custody associated with evidence "
            f"{ev_id}. Custody records belonging to "
            f"other evidence items are excluded."
        )

        subject = ev_id

    # ──────────────────────────────────────────────────────────
    # AUDIT REPORT
    # ──────────────────────────────────────────────────────────

    elif report_type == "audit":

        audits = AuditLog.all(
            db,
            limit=1000
        )

        add_header(
            "System Audit Report",
            (
                f"EvidenceChain Audit Trail | "
                f"Generated {now_str} by "
                f"{current_user.name}"
            )
        )

        add_section(
            "1. Automatically Generated Audit Summary"
        )

        add_text(
            generate_audit_summary(
                audits
            )
        )

        add_section(
            "2. Audit Events"
        )

        if audits:

            add_table(
                [
                    "Timestamp",
                    "Actor",
                    "Action",
                    "Module",
                    "Target",
                    "Detail"
                ],
                [
                    [
                        _fmt_dt(
                            a.get("timestamp")
                        ),
                        a.get("actor"),
                        a.get("action"),
                        a.get("module"),
                        a.get("target"),
                        a.get("detail")
                    ]
                    for a in audits
                ],
                [
                    2.3 * cm,
                    2.2 * cm,
                    2.2 * cm,
                    1.8 * cm,
                    2.3 * cm,
                    6.0 * cm
                ]
            )

        else:

            add_text(
                "No audit events are currently recorded."
            )

        add_section(
            "3. Audit Conclusion"
        )

        add_text(
            "This report is restricted to the "
            "EvidenceChain system audit trail. "
            "The underlying audit records are "
            "not modified during report generation."
        )

        subject = "audit"

    else:

        abort(
            404,
            "Unknown report type."
        )

    return (
        elements,
        subject,
        A4,
        cm,
        colors
    )


# ──────────────────────────────────────────────────────────────
# PDF Export
# ──────────────────────────────────────────────────────────────

@reports_bp.route(
    "/export/pdf/<report_type>"
)
@login_required
def export_pdf(report_type):

    db = _db()

    now = datetime.now(
        timezone.utc
    )

    now_str = now.strftime(
        "%Y-%m-%d %H:%M UTC"
    )

    try:

        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate
        )

    except ImportError:

        abort(
            500,
            "ReportLab is not installed."
        )

    (
        elements,
        subject,
        A4,
        cm,
        colors
    ) = _build_pdf(
        report_type,
        db,
        now_str
    )

    buffer = io.BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=1.7 * cm,
        bottomMargin=1.8 * cm,
        title=(
            f"EvidenceChain "
            f"{report_type} report"
        ),
        author=current_user.name
    )

    def footer(
        canvas,
        doc
    ):

        canvas.saveState()

        canvas.setStrokeColor(
            colors.HexColor(
                "#cbd5e1"
            )
        )

        canvas.line(
            2 * cm,
            1.35 * cm,
            19 * cm,
            1.35 * cm
        )

        canvas.setFont(
            "Helvetica",
            7.5
        )

        canvas.setFillColor(
            colors.HexColor(
                "#64748b"
            )
        )

        canvas.drawString(
            2 * cm,
            0.85 * cm,
            "EvidenceChain Digital Evidence Management System"
        )

        canvas.drawRightString(
            19 * cm,
            0.85 * cm,
            f"Page {doc.page}"
        )

        canvas.restoreState()

    document.build(
        elements,
        onFirstPage=footer,
        onLaterPages=footer
    )

    buffer.seek(0)

    filename = (
        f"evidencechain_"
        f"{report_type}_"
        f"{_safe_filename(subject)}_"
        f"{now.strftime('%Y%m%d')}.pdf"
    )

    return send_file(
        buffer,
        mimetype="application/pdf",
        download_name=filename,
        as_attachment=True
    )


# ──────────────────────────────────────────────────────────────
# Excel Export
# ──────────────────────────────────────────────────────────────

@reports_bp.route(
    "/export/excel/<report_type>"
)
@login_required
def export_excel(report_type):

    db = _db()

    try:

        import openpyxl

        from openpyxl.styles import (
            Font,
            PatternFill,
            Alignment
        )

    except ImportError:

        abort(
            500,
            "openpyxl is not installed."
        )

    now = datetime.now(
        timezone.utc
    )

    now_str = now.strftime(
        "%Y-%m-%d %H:%M UTC"
    )

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.sheet_view.showGridLines = False

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="0F172A"
    )

    def headers(values):

        sheet.append(values)

        for cell in sheet[1]:

            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    # ──────────────────────────────────────────────────────────
    # Case
    # ──────────────────────────────────────────────────────────

    if report_type == "cases":

        case_id = request.args.get(
            "case_id",
            ""
        ).strip()

        if not case_id:
            abort(
                400,
                "case_id is required."
            )

        case = Case.get(
            db,
            case_id
        )

        if not case:
            abort(
                404,
                "Case not found."
            )

        evidence = Evidence.by_case(
            db,
            case_id
        )

        sheet.title = "Case Report"

        headers([
            "Field",
            "Value"
        ])

        sheet.append([
            "Report Type",
            "Digital Case Report"
        ])

        sheet.append([
            "Generated By",
            current_user.name
        ])

        sheet.append([
            "Generated At",
            now_str
        ])

        sheet.append([
            "Case ID",
            case.get("case_id")
        ])

        sheet.append([
            "Case Title",
            case.get("title")
        ])

        sheet.append([
            "Case Type",
            case.get("type")
        ])

        sheet.append([
            "Status",
            case.get("status")
        ])

        sheet.append([
            "Priority",
            case.get("priority")
        ])

        sheet.append([
            "Lead Investigator",
            case.get("lead")
        ])

        sheet.append([
            "Location",
            case.get("location")
        ])

        sheet.append([
            "Opened",
            _fmt_dt(
                case.get("created")
            )
        ])

        sheet.append([
            "Generated Description",
            generate_case_description(
                case,
                len(evidence)
            )
        ])

        sheet.append([])

        sheet.append([
            "Linked Evidence"
        ])

        sheet.append([
            "Evidence ID",
            "Name",
            "Type",
            "Status",
            "Collected By",
            "Collected At"
        ])

        for e in evidence:

            sheet.append([
                e.get("evidence_id"),
                e.get("name"),
                e.get("type"),
                e.get("status"),
                e.get("collected_by"),
                _fmt_dt(
                    e.get("collected_at")
                )
            ])

        subject = case_id

    # ──────────────────────────────────────────────────────────
    # Evidence
    # ──────────────────────────────────────────────────────────

    elif report_type == "evidence":

        ev_id = request.args.get(
            "ev_id",
            ""
        ).strip()

        if not ev_id:
            abort(
                400,
                "ev_id is required."
            )

        ev = Evidence.get(
            db,
            ev_id
        )

        if not ev:
            abort(
                404,
                "Evidence not found."
            )

        case = Case.get(
            db,
            ev.get("case_id", "")
        )

        custody = CustodyLog.by_evidence(
            db,
            ev_id
        )

        sheet.title = "Evidence Report"

        headers([
            "Field",
            "Value"
        ])

        rows = [
            (
                "Report Type",
                "Digital Evidence Report"
            ),
            (
                "Generated By",
                current_user.name
            ),
            (
                "Generated At",
                now_str
            ),
            (
                "Evidence ID",
                ev.get("evidence_id")
            ),
            (
                "Evidence Name",
                ev.get("name")
            ),
            (
                "Evidence Type",
                ev.get("type")
            ),
            (
                "Status",
                ev.get("status")
            ),
            (
                "Linked Case",
                ev.get("case_id")
            ),
            (
                "Case Title",
                case.get("title")
                if case
                else "Not found"
            ),
            (
                "Original Filename",
                ev.get("filename")
            ),
            (
                "File Size",
                ev.get("file_size")
            ),
            (
                "Storage Location",
                ev.get("location")
            ),
            (
                "Collected By",
                ev.get("collected_by")
            ),
            (
                "Collected At",
                _fmt_dt(
                    ev.get("collected_at")
                )
            ),
            (
                "SHA-256",
                ev.get("hash")
            ),
            (
                "Collection Notes",
                ev.get("notes")
            ),
            (
                "Generated Description",
                generate_evidence_description(
                    ev,
                    case
                )
            )
        ]

        for key, value in rows:

            sheet.append([
                key,
                value
            ])

        sheet.append([])

        sheet.append([
            "Chain of Custody"
        ])

        sheet.append([
            "Timestamp",
            "Action",
            "From",
            "To",
            "Handler",
            "Reason",
            "Location",
            "Integrity"
        ])

        for c in custody:

            sheet.append([
                _fmt_dt(
                    c.get("timestamp")
                ),
                c.get("action"),
                c.get("from_party"),
                c.get("to_party"),
                c.get("handler"),
                c.get("reason"),
                c.get("location"),
                c.get("integrity")
            ])

        subject = ev_id

    # ──────────────────────────────────────────────────────────
    # Custody
    # ──────────────────────────────────────────────────────────

    elif report_type == "custody":

        ev_id = request.args.get(
            "ev_id",
            ""
        ).strip()

        if not ev_id:
            abort(
                400,
                "ev_id is required."
            )

        ev = Evidence.get(
            db,
            ev_id
        )

        if not ev:
            abort(
                404,
                "Evidence not found."
            )

        logs = CustodyLog.by_evidence(
            db,
            ev_id
        )

        sheet.title = "Custody Report"

        headers([
            "Field",
            "Value"
        ])

        rows = [
            (
                "Report Type",
                "Chain of Custody Report"
            ),
            (
                "Generated By",
                current_user.name
            ),
            (
                "Generated At",
                now_str
            ),
            (
                "Evidence ID",
                ev.get("evidence_id")
            ),
            (
                "Evidence Name",
                ev.get("name")
            ),
            (
                "Evidence Type",
                ev.get("type")
            ),
            (
                "Evidence Status",
                ev.get("status")
            ),
            (
                "SHA-256",
                ev.get("hash")
            ),
            (
                "Custody Event Count",
                len(logs)
            ),
            (
                "Generated Summary",
                generate_custody_summary(
                    ev,
                    logs
                )
            )
        ]

        for key, value in rows:

            sheet.append([
                key,
                value
            ])

        sheet.append([])

        sheet.append([
            "Custody Timeline"
        ])

        sheet.append([
            "Timestamp",
            "Action",
            "From",
            "To",
            "Handler",
            "Reason",
            "Location",
            "Integrity"
        ])

        for c in logs:

            sheet.append([
                _fmt_dt(
                    c.get("timestamp")
                ),
                c.get("action"),
                c.get("from_party"),
                c.get("to_party"),
                c.get("handler"),
                c.get("reason"),
                c.get("location"),
                c.get("integrity")
            ])

        subject = ev_id

    # ──────────────────────────────────────────────────────────
    # Audit
    # ──────────────────────────────────────────────────────────

    elif report_type == "audit":

        audits = AuditLog.all(
            db,
            limit=5000
        )

        sheet.title = "Audit Report"

        headers([
            "Timestamp",
            "Actor",
            "Action",
            "Module",
            "Target",
            "Detail",
            "IP"
        ])

        for a in audits:

            sheet.append([
                _fmt_dt(
                    a.get("timestamp")
                ),
                a.get("actor"),
                a.get("action"),
                a.get("module"),
                a.get("target"),
                a.get("detail"),
                a.get("ip")
            ])

        subject = "audit"

    else:

        abort(
            404,
            "Unknown report type."
        )

    # ──────────────────────────────────────────────────────────
    # Excel formatting
    # ──────────────────────────────────────────────────────────

    for column in sheet.columns:

        maximum = 0

        for cell in column:

            value = str(
                cell.value
                or ""
            )

            maximum = max(
                maximum,
                len(value)
            )

        sheet.column_dimensions[
            column[0].column_letter
        ].width = min(
            max(maximum + 3, 12),
            60
        )

    for row in sheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    buffer = io.BytesIO()

    workbook.save(
        buffer
    )

    buffer.seek(0)

    filename = (
        f"evidencechain_"
        f"{report_type}_"
        f"{_safe_filename(subject)}_"
        f"{now.strftime('%Y%m%d')}.xlsx"
    )

    return send_file(
        buffer,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        download_name=filename,
        as_attachment=True
    )